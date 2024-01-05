import requests
import pandas as pd
import pyperclip
import time
import openpyxl

#from requests_html import HTMLSession
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.edge.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC





def getAppDecriptions(driver, url1):
    #print(url1)
    driver.get(url1)
    showDes = 'showDescription(); return false' #comando para expandir a descricao
    driver.execute_script(showDes)
    #showTec = 'clickShowTech()'     #comando expandir as tecnologias usadas
    #driver.execute_script(showTec)
    shortDes = 'ERROR!*!'
    aboutApp = 'ERROR!*!'
    fulllDes = 'ERROR!*!'
    tableApp = 'ERROR!*!'
    a = 1
    time.sleep(1)
    try:
        shortDes = driver.find_element(By.CLASS_NAME, 'app-short-description').text #app short description
    except Exception as e:
        a = a + 1
    try:
        aboutApp = driver.find_elements(By.CLASS_NAME, 'col-12')[0].text #about the application
    except Exception as e:
        a = a + 1
    try:
        fulllDes = driver.find_elements(By.CLASS_NAME, 'col-12')[1].text #full description
    except Exception as e:
        a = a + 1
    try:
        tableApp = driver.find_elements(By.CLASS_NAME, 'col-12')[2].text #table of caracteristics
    except Exception as e:
        a = a + 1
    
    #print(driver.find_element(By.ID, 'descContents').text) #printa a descricao
    #print('================')
    #print(driver.find_element(By.CLASS_NAME, 'app-short-description').text) #printa a descricao curta
    #print(driver.find_elements(By.CLASS_NAME, 'col-12')[0].text) #printa o about
    #print(driver.find_elements(By.CLASS_NAME, 'col-12')[1].text) #printa a descricao
    #print(driver.find_elements(By.CLASS_NAME, 'col-12')[2].text) #printa,se tiver, a tabela

    return [shortDes, aboutApp, fulllDes,tableApp]



def getAppTecRankCom(driver, url1):
    #print(url1)
    driver.get(url1)
    showLog = 'clickShowChangelog()' #comando para mostrar o changelog
    driver.execute_script(showLog)
    showTec = 'clickShowTech()'     #comando expandir as tecnologias usadas
    driver.execute_script(showTec)
    shortDes = 'ERROR!*!'
    aboutApp = 'ERROR!*!'
    fulllDes = 'ERROR!*!'
    tableApp = 'ERROR!*!'
    a = 1
    #time.sleep(9) #timer de teste
    time.sleep(1) #timer para coleta das informações da página
    try:
        shortDes = driver.find_element(By.CLASS_NAME, 'data-table-container').text #tabela ranking
    except Exception as e:
        a = a + 1
    try:
        aboutApp = driver.find_elements(By.CLASS_NAME, 'mb-1')
        comments = ''
        for i in aboutApp:
            comments = comments + '\n' + i.text
    except Exception as e:
        a = a + 1
    try:
        fulllDes = driver.find_element(By.ID, 'techContents').text #technologies
    except Exception as e:
        a = a + 1
    try:
        tableApp = driver.find_element(By.CLASS_NAME, 'app-changelog').text #changelog
    except Exception as e:
        a = a + 1
    
    print('================')
    print(shortDes)
    print(comments)
    print(fulllDes)
    print(tableApp)
    #print(driver.find_element(By.ID, 'descContents').text) #printa a descricao
    #print('================')
    #print(driver.find_element(By.CLASS_NAME, 'app-short-description').text) #printa a descricao curta
    #print(driver.find_elements(By.CLASS_NAME, 'col-12')[0].text) #printa o about
    #print(driver.find_elements(By.CLASS_NAME, 'col-12')[1].text) #printa a descricao
    #print(driver.find_elements(By.CLASS_NAME, 'col-12')[2].text) #printa,se tiver, a tabela

    return [shortDes, comments, fulllDes, tableApp]

def getAlllllllll(driver, url1):
    #print(url1)
    driver.get(url1)
    showDes = 'showDescription(); return false' #comando para expandir a descricao
    driver.execute_script(showDes)
    showLog = 'clickShowChangelog()' #comando para mostrar o changelog
    driver.execute_script(showLog)
    showTec = 'clickShowTech()'     #comando expandir as tecnologias usadas
    try:
        driver.execute_script(showTec)
    except Exception as e:
        print("NoTech")
    


    infos = []
    for i in range(9):
        infos.append('ERROR!*!')
    a = 1
    #time.sleep(9) #timer de teste
    time.sleep(1) #timer para coleta das informações da página
    try:
        infos[0] = driver.find_element(By.CLASS_NAME, 'data-table-container').text #tabela ranking
    except Exception as e:
        a = a + 1
    try:
        aboutApp = driver.find_elements(By.CLASS_NAME, 'mb-1')
        infos[1] = ''
        for i in aboutApp:
            infos[1] = infos[1] + '\n' + i.text
    except Exception as e:
        a = a + 1
    try:
        infos[2] = driver.find_element(By.ID, 'techContents').text #technologies
    except Exception as e:
        a = a + 1
    try:
        infos[3] = driver.find_element(By.CLASS_NAME, 'app-changelog').text #changelog
    except Exception as e:
        a = a + 1
    try:
        infos[4] = driver.find_element(By.CLASS_NAME, 'app-short-description').text #app short description
    except Exception as e:
        a = a + 1
    try:
        infos[5] = driver.find_elements(By.CLASS_NAME, 'col-12')[0].text #about the application
    except Exception as e:
        a = a + 1
    try:
        infos[6] = driver.find_elements(By.CLASS_NAME, 'col-12')[1].text #full description
    except Exception as e:
        a = a + 1
    try:
        infos[7] = driver.find_elements(By.CLASS_NAME, 'col-12')[2].text #table of caracteristics
    except Exception as e:
        a = a + 1
    try:
        infos[8] = driver.find_element(By.CLASS_NAME, 'app-top-title').text #appName
    except Exception as e:
        a = a + 1


    
    #print('================')
    #print(shortDes)
    #print(comments)
    #print(fulllDes)
    #print(tableApp)
    #print(driver.find_element(By.ID, 'descContents').text) #printa a descricao
    #print('================')
    #print(driver.find_element(By.CLASS_NAME, 'app-short-description').text) #printa a descricao curta
    #print(driver.find_elements(By.CLASS_NAME, 'col-12')[0].text) #printa o about
    #print(driver.find_elements(By.CLASS_NAME, 'col-12')[1].text) #printa a descricao
    #print(driver.find_elements(By.CLASS_NAME, 'col-12')[2].text) #printa,se tiver, a tabela

    return infos


def listaInfo():
    service = Service()
    options = webdriver.EdgeOptions()
    #options.add_argument("--headless") #não renderiza interface gráfica
    driver = webdriver.Edge(service=service, options=options)
    # Nome do arquivo Excel a ser lido e atualizado
    path = "F:\\Projetinhos\\Projeto Dissertação\\PythonColetaDados\\"
    nome = ""
    extension = ".xlsx"

    linhaInicial = 0
    finalValue = 0

    file = 2

    if(file == 1):
        nome = "totalInfo2"
        linhaInicial =  2
        finalValue = 1600
    if(file == 2):
        nome = "totalInfo1600"
        linhaInicial =  1600
        finalValue = 3200
    if(file == 3):
        nome = "totalInfo3200"
        linhaInicial =  3200
        finalValue = 4800
    if(file == 4):
        nome = "totalInfo4800"
        linhaInicial =  4800
        finalValue = 7000  
    
    linhaInicial = 2219

    countLinha = linhaInicial
    savePoint = countLinha + 400




    #inicialmente ele abre um app com acesso restrito, apenas para deixar sinalizado que quero acessar esses apps
    driver.get('https://www.appbrain.com/app/beechat-dating-nearby/th.cyberapp.beechat')
    time.sleep(5)

    # Abre o arquivo Excel
    workbook = openpyxl.load_workbook(path + nome + extension)
    sheet = workbook.active

    # Itera pelas linhas na primeira coluna do arquivo Excel
    for row in sheet.iter_rows(min_row=linhaInicial, max_row=sheet.max_row, min_col=1, max_col=1):
        if countLinha == finalValue: break
        for cell in row:
            texto = cell.value
            if texto is not None:
                # Obtém os valores da função processarTexto
                #valores = getAppDecriptions(driver, texto) # quando eu estava pegando as descrições
                #valores = getPlayStore(driver, texto)
                valores = getAlllllllll(driver, texto)
                
                # Atualiza as colunas 2 a 5 com os valores retornados
                for i, val in enumerate(valores, start=3):
                    sheet.cell(row=cell.row, column=i, value=val)

                if countLinha == savePoint:
                    workbook.save(path + nome + '-' + str(countLinha) + extension)
                    savePoint = countLinha + 400
                workbook.save(path + nome + extension)
                countLinha = countLinha + 1
                print(countLinha)

    # Salva as alterações de volta no arquivo Excel
    

    print("Coleta finalizada.")


listaInfo()

