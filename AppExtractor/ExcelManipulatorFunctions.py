import requests
import pandas as pd
import pyperclip
import time
import openpyxl

#from requests_html import HTMLSession
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.edge.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC



def setPlayStore(url1):

    diretorio = ""
    playstore = ""
    partes = url1.split('/')
    # Verifica se existem pelo menos quatro ocorrências de "/"
    if len(partes) >= 6:
        # Retorna o texto após a quarta ocorrência
        diretorio = '/'.join(partes[5:])
        playstore = 'https://play.google.com/store/apps/details?id=' + diretorio
        print(diretorio)
    else:
        # Se não houver quatro ocorrências, retorne uma string vazia ou outra mensagem de erro, conforme necessário
        print("Erro")

    return [diretorio, playstore]

def extrair_caracteristicas(string):
    # Lista de características
    caracteristicas = [
        "More data about",
        "Price",
        "Cost of in-app purchases",
        "Total downloads",
        "Recent downloads",
        "Rating",
        "Ranking",
        "Version",
        "APK size",
        "Number of libraries",
        "Designed for",
        "Maturity",
        "Suitable for",
        "Ads"
    ]

    # Inicializa o vetor de valores com "N/A"
    valores = ["N/A"] * 14

    # Divide a string em linhas
    linhas = string.split('\n')
    
    # Para cada linha, verifica se corresponde a alguma característica e extrai o valor
    for linha in linhas:
        for i, caracteristica in enumerate(caracteristicas):
            if linha.startswith(caracteristica):
                # Extrai o valor removendo a parte da característica
                valores[i] = linha[len(caracteristica):].strip()

    return valores

def converter_para_numero(string):
    if(string == "") : return [""]
    # Mapeia os sufixos para os fatores correspondentes
    sufixos = {'billion': 1_000_000_000, 'million': 1_000_000, 'thousand': 1_000}

    # Divide a string em partes (número e sufixo, se existir)
    partes = string.split()

    # Converte a parte numérica para float
    numero_str = partes[0].replace(',', '')  # Remove vírgulas, se houver
    numero = float(numero_str)

    # Multiplica pelo fator correspondente ao sufixo, se existir
    if len(partes) > 1:
        numero *= sufixos[partes[1]]

    return [str(int(numero))]

def separate_tech(string):

    # Divide a string em linhas
    linhas = string.split('\n')

    tech = False

    strPerm = ""
    strTech = ""
    
    # Para cada linha, verifica se corresponde a alguma característica e extrai o valor
    for linha in linhas:
        if tech:
            strTech = strTech + linha + "\n"
        else : 
            if linha.startswith("Technologies used by "):
                tech = True
            else: 
                strPerm = strPerm + linha + "\n"
    
    if len(strTech) == 0:
        strTech = "N/A"

    if len(strPerm) == 0:
        strPerm = "N/A"
    
    return [strPerm, strTech]


def rating(string):
    valores = []

    if string != 'No ratings':
        valores = string.split(' based on ')
        valores[1] = valores[1].replace(' ratings', '')

    return valores

def noAndroid(string):
    text =  []
    if string.startswith('Android '):
        text.append(string.replace('Android ', ''))
        text[0] = text[0].replace('+', '')
    return text

def noPerm(string):
    string[0]
    if string.startswith('N/A'):
        return string
    else:
        string

def remover_primeira_linha(texto):
    linhas = texto.split('Description\n', 1)
    if len(linhas) > 1:
        return [linhas[1]]
    else:
        return ['N/A']


def processar_linhas_pares(texto, lista_geral):
    linhas = texto.split('\n')
    
    # Adiciona o texto das linhas pares à lista geral
    for i in range(0, len(linhas), 2):
        texto_linha_par = linhas[i].strip()  # Remove espaços em branco no início e no final
        if texto_linha_par not in lista_geral:
            lista_geral.append(texto_linha_par)


def processar_string(string, lista_fixa):
    # Inicializa a lista com "N/A"
    lista_resultado = ["N/A"] * 12

    # Divide a string em linhas
    linhas = string.split('\n')

    # Itera pelas linhas pares
    for i in range(0, len(linhas), 2):
        texto_linha_par = linhas[i].strip()  # Remove espaços em branco no início e no final

        # Verifica se o texto da linha par corresponde a algum elemento da lista fixa
        if texto_linha_par in lista_fixa:
            # Encontra a posição na lista fixa
            posicao = lista_fixa.index(texto_linha_par)

            # Se a próxima linha existe (linha ímpar), atribui à lista resultado
            if i + 1 < len(linhas):
                lista_resultado[posicao] = linhas[i + 1].strip()

    return lista_resultado

import re
def extrair_trechos(texto):
    # Expressões regulares para os dois trechos de texto
    padrao1 = re.compile(r' is (.*?)app developed by ')
    padrao2 = re.compile(r'app developed by (.*?)\. The APK')

    # Procura pelos padrões no texto
    resultado1 = re.search(padrao1, texto)
    resultado2 = re.search(padrao2, texto)

    # Extrai os trechos correspondentes
    trecho1 = resultado1.group(1) if resultado1 else None
    trecho2 = resultado2.group(1) if resultado2 else None

    return [trecho1, trecho2]


def formatar_texto(texto):
    # Lista de prefixos a serem removidos
    prefixos = ["a ", "an "]

    # Verifica se o texto possui algum dos prefixos
    for prefixo in prefixos:
        if texto.startswith(prefixo):
            # Remove o prefixo e capitaliza a primeira letra de cada palavra
            palavras = texto[len(prefixo):].split()
            texto_formatado = ' '.join(word.capitalize() for word in palavras)
            return [texto_formatado]

    # Se nenhum prefixo for encontrado, retorna o texto original
    return [texto]

def categorizar_linhas(string):
    positivos = ""
    negativos = ""

    # Separa as linhas
    linhas = string.split('\n')

    for linha in linhas:
        if linha.startswith('★★☆☆☆ '):
            if negativos != "":
                negativos = negativos + "\n"
            negativos = negativos + linha.replace('★★☆☆☆ ', '')
        elif linha.startswith('★★★★★ '):
            if positivos != "":
                positivos = positivos + "\n"
            positivos = positivos + linha.replace('★★★★★ ', '')

    # Se as listas estiverem vazias, atribui "N/A"
    if positivos == "":
        positivos = "N/A"
    if negativos  == "":
        negativos = "N/A"

    return [positivos, negativos]

def tirarVirgulas(texto):
    return [texto.replace(', ', '\n')]

def offOrganizer(selected, writed):

    #service = Service()
    #options = webdriver.EdgeOptions()
    #options.add_argument("--headless") #não renderiza interface gráfica
    #driver = webdriver.Edge(service=service, options=options)
    # Nome do arquivo Excel a ser lido e atualizado
    path = "F:\\Projetinhos\\Projeto Dissertação\\PythonColetaDados\\"
    nome = "totalInfoComplete"
    extension = ".xlsx"

    linhaInicial = 3
    countLinha = linhaInicial
    
    finalValue = 1

    #coluna que o texto será selecionado
    selectCol = selected
    #primeira coluna que será escrita
    firtWritingCol = writed




    #inicialmente ele abre um app com acesso restrito, apenas para deixar sinalizado que quero acessar esses apps
    #driver.get('https://www.appbrain.com/app/beechat-dating-nearby/th.cyberapp.beechat')
    #time.sleep(5)

    # Abre o arquivo Excel
    workbook = openpyxl.load_workbook(path + nome + extension)
    sheet = workbook.active

    #lista_geral = ['Development tools', 'Network communication', 'Storage', 'Hardware controls', 'System tools',  'Your accounts', 'Your location', 'Your personal information', 'Phone calls', 'Services that cost you money', 'Your messages', 'Extra']

    # Itera pelas linhas na primeira coluna do arquivo Excel
    for row in sheet.iter_rows(min_row=linhaInicial, max_row=sheet.max_row, min_col=selectCol, max_col=selectCol):
        if countLinha == finalValue: break
        for cell in row:
            texto = cell.value
            countLinha = countLinha + 1
            #print(texto)
            if texto is not None:
                # Obtém os valores da função processarTexto
                #valores = extrair_caracteristicas(texto)
                #valores = remover_primeira_linha(texto)
                #valores = processar_string(texto, lista_geral)

                #valores = formatar_texto(texto)
                #valores = categorizar_linhas(texto)
                valores = converter_para_numero(texto)
                #looping para adicionar os elementos em valores às colunas da tabela
                for i, val in enumerate(valores, start=firtWritingCol):
                    sheet.cell(row=cell.row, column=i, value=val)

                #print(countLinha)

    # Salva as alterações de volta no arquivo Excel
    workbook.save(path + nome + extension)

    print("Organização finalizada.")

offOrganizer(15,50)
offOrganizer(16,51)
offOrganizer(26,52)








# Exemplo de uso converter_para_numero(string)
#strings_exemplo = ["1.1 billion", "12 million", "7.6 million", "230 thousand", "370", ""]
#for string in strings_exemplo:
#    resultado = converter_para_numero(string)[0]
#    print(f"{string}: {resultado}")
    



#print(tirarVirgulas("""approximate (network-based) location, precise (GPS) location""")[0])

