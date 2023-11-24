import requests
import pandas as pd
import pyperclip
import time
import openpyxl

#from requests_html import HTMLSession
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.edge.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC



def vamoBR():
    service = Service()
    options = webdriver.EdgeOptions()
    driver = webdriver.Edge(service=service, options=options)

    url = 'https://www.appbrain.com/app/canva-design-photo-video/com.canva.editor'

    driver.get(url)

    #clicka nos elementos para carregar as informações dinâmicas da tela.
    #driver.find_element(By.ID, 'descLink').click()
    #driver.find_element(By.CLASS_NAME, 'link').click()
    #driver.execute('clickShowTech()')
    appName         = driver.find_element(By.CLASS_NAME, 'app-top-title').text
    #appDescription  = driver.find_element(By.ID, 'descContents').text
    appDeveloper    = driver.find_element(By.XPATH, '//*[@id="main_content"]/div[1]/div/div[2]/span').text
    appAbout        = driver.find_elements(By.CLASS_NAME, 'table table-striped')


    print("\n\n\n\n\n\n\n")
    print(appName)
    #print(appDescription)
    #print(appDeveloper)
    print(appAbout.text)

def beatifulSoupSeleniumWire():
    #service = Service()
    #options = webdriver.EdgeOptions()
    #options.add_argument("--headless") #não renderiza interface gráfica
    driver = webdriver.Chrome()

    driver.get('https://www.appbrain.com/app/wonder-ai-art-generator/com.codeway.wonder')
    showDes = 'showDescription(); return false' #comando para expandir a descricao
    showTec = 'clickShowTech()'     #comando expandir as tecnologias usadas
    driver.execute_script(showDes)
    driver.execute_script(showTec)
    #driver.get('https://www.appbrain.com/app/wonder-ai-art-generator/com.codeway.wonder')
    
    time.sleep(0.5)
    site = BeautifulSoup(driver.page_source, "html.parser")
    #time.sleep(5)

    #print(site.prettify())
    findSpan = site.find_all("span")
    findTd = site.find_all("td")

    #print(findTd)
    count = 0
    for x in findTd:
        print(str(count) + ' - ' + str(x))
        count = count + 1


    driver.close

    #header = ['appName','developerName','ratingValue','ratingCount','downloads','shortDescription','Cost of in-app purchases','Total downloads','Recent downloads','Rating','Ranking','Version','APK size',
    #          'Number of libraries','Designed for Android','Suitable for','Ads',]
    #app = []
    #app.append(driver.find_element(By.XPATH, '/html/body/div[2]/div[3]/div[1]/div[1]/div[2]/div[1]/div/div[2]/div[1]/h1').text)
    #app.append(driver.find_element(By.XPATH, '/html/body/div[2]/div[3]/div[1]/div[1]/div[2]/div[1]/div/div[2]/span/a/span').text)
    #app.append(driver.find_element(By.XPATH, '/html[1]/body[1]/div[2]/div[3]/div[1]/div[1]/div[2]/div[2]/div[1]/span[1]').text)
    #app.append(driver.find_element(By.XPATH, '/html[1]/body[1]/div[2]/div[3]/div[1]/div[1]/div[2]/div[2]/div[2]/span[1]').text)
    #app.append(driver.find_element(By.XPATH, '/html/body/div[2]/div[3]/div[1]/div[1]/div[2]/div[2]/div[3]').text)
    #app.append(driver.find_element(By.XPATH, '/html/body/div[2]/div[3]/div[1]/div[1]/div[2]/h2[1]').text)
    #
    #print(app)

def beatifulSoupSelenium():
    service = Service()
    options = webdriver.EdgeOptions()
    #options.add_argument("--headless") #não renderiza interface gráfica
    driver = webdriver.Edge(service=service, options=options)

    driver.get('https://www.appbrain.com/app/wonder-ai-art-generator/com.codeway.wonder')
    showDes = 'showDescription(); return false' #comando para expandir a descricao
    showTec = 'clickShowTech()'     #comando expandir as tecnologias usadas
    driver.execute_script(showDes)
    #driver.execute_script(showTec)

    #print(driver.find_element(By.ID, 'descContents').text) #printa a descricao
    print('================')
    #print(driver.find_element(By.CLASS_NAME, 'app-short-description').text) #printa a descricao curta
    #print(driver.find_elements(By.CLASS_NAME, 'col-12')[0].text) #printa o about
    #print(driver.find_elements(By.CLASS_NAME, 'col-12')[1].text) #printa a descricao
    #print(driver.find_elements(By.CLASS_NAME, 'col-12')[2].text) #printa,se tiver, a tabela
    #print(driver.find_elements(By.ID, 'techContents')[0].text)
    
    #time.sleep(0.5)
    #site = BeautifulSoup(driver.page_source, "html.parser")
    #time.sleep(5)

    #print(site.prettify())
    #findSpan = site.find("span", itemprop="ratingValue")
    #findTd = site.find_all("td")
    findSpan = driver.find_elements(By.ID, 'techContents')
    #print(findSpan)
    count = 0
    for x in findSpan:
        print(str(count) + ' - ' + str(x.text))
        count = count + 1


    driver.close

    #header = ['appName','developerName','ratingValue','ratingCount','downloads','shortDescription','Cost of in-app purchases','Total downloads','Recent downloads','Rating','Ranking','Version','APK size',
    #          'Number of libraries','Designed for Android','Suitable for','Ads',]
    #app = []
    #app.append(driver.find_element(By.XPATH, '/html/body/div[2]/div[3]/div[1]/div[1]/div[2]/div[1]/div/div[2]/div[1]/h1').text)
    #app.append(driver.find_element(By.XPATH, '/html/body/div[2]/div[3]/div[1]/div[1]/div[2]/div[1]/div/div[2]/span/a/span').text)
    #app.append(driver.find_element(By.XPATH, '/html[1]/body[1]/div[2]/div[3]/div[1]/div[1]/div[2]/div[2]/div[1]/span[1]').text)
    #app.append(driver.find_element(By.XPATH, '/html[1]/body[1]/div[2]/div[3]/div[1]/div[1]/div[2]/div[2]/div[2]/span[1]').text)
    #app.append(driver.find_element(By.XPATH, '/html/body/div[2]/div[3]/div[1]/div[1]/div[2]/div[2]/div[3]').text)
    #app.append(driver.find_element(By.XPATH, '/html/body/div[2]/div[3]/div[1]/div[1]/div[2]/h2[1]').text)
    #
    #print(app)

def getInfo(driver, url1):
    #print(url1)
    driver.get(url1)
    showDes = 'showDescription(); return false' #comando para expandir a descricao
    driver.execute_script(showDes)
    #showTec = 'clickShowTech()'     #comando expandir as tecnologias usadas
    #driver.execute_script(showTec)
    shortDes = 'ERROR!*!'
    aboutApp = 'ERROR!*!'
    fulllDes = 'ERROR!*!'
    tableApp = 'ERROR!*!'
    a = 1
    #time.sleep(2)
    try:
        shortDes = driver.find_element(By.CLASS_NAME, 'app-short-description').text
    except Exception as e:
        a = a + 1
    try:
        aboutApp = driver.find_elements(By.CLASS_NAME, 'col-12')[0].text
    except Exception as e:
        a = a + 1
    try:
        fulllDes = driver.find_elements(By.CLASS_NAME, 'col-12')[1].text
    except Exception as e:
        a = a + 1
    try:
        tableApp = driver.find_elements(By.CLASS_NAME, 'col-12')[2].text
    except Exception as e:
        a = a + 1
    
    #print(driver.find_element(By.ID, 'descContents').text) #printa a descricao
    #print('================')
    #print(driver.find_element(By.CLASS_NAME, 'app-short-description').text) #printa a descricao curta
    #print(driver.find_elements(By.CLASS_NAME, 'col-12')[0].text) #printa o about
    #print(driver.find_elements(By.CLASS_NAME, 'col-12')[1].text) #printa a descricao
    #print(driver.find_elements(By.CLASS_NAME, 'col-12')[2].text) #printa,se tiver, a tabela

    return [shortDes, aboutApp, fulllDes,tableApp]


def listaInfo():
    service = Service()
    options = webdriver.EdgeOptions()
    #options.add_argument("--headless") #não renderiza interface gráfica
    driver = webdriver.Edge(service=service, options=options)
    # Nome do arquivo Excel a ser lido e atualizado
    arquivo_excel = "exemplo.xlsx"

    # Abre o arquivo Excel
    workbook = openpyxl.load_workbook(arquivo_excel)
    sheet = workbook.active

    # Itera pelas linhas na primeira coluna do arquivo Excel
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
        for cell in row:
            texto = cell.value
            if texto is not None:
                # Obtém os valores da função processarTexto
                valores = getInfo(driver, texto)
                # Atualiza as colunas 2 a 5 com os valores retornados
                for i, val in enumerate(valores, start=2):
                    sheet.cell(row=cell.row, column=i, value=val)

    # Salva as alterações de volta no arquivo Excel
    workbook.save(arquivo_excel)

    print("Planilha atualizada e salva com sucesso.")

def bot_lindinho_2009():
    service = Service()
    options = webdriver.EdgeOptions()
    #options.add_argument("--headless") #não renderiza interface gráfica
    driver = webdriver.Edge(service=service, options=options)
    driver.get('')

    try:
        temp = driver.find_elements(By.CLASS_NAME, 'btComDet')
        print(temp[1].text)
    except Exception as e:
        print(e)
    
    time.sleep(30000)


def semZeradosPfv():
    service = Service()
    options = webdriver.EdgeOptions()
    driver = webdriver.Edge(service=service, options=options)

    lista = ['https://www.appbrain.com/app/blk-dating-meet-black-singles/com.affinityapps.blk','https://www.appbrain.com/app/y-hookup-app-fwb-adult-dating/com.tabooapp.dating','https://www.appbrain.com/app/porn-blocker-safe-search/com.kawkaw.pornblocker.safebrowser.up','https://www.appbrain.com/app/peach-live/com.peach.live','https://www.appbrain.com/app/face-chat/com.facechat.live','https://www.appbrain.com/app/private-safe-browser/io.happybrowsing','https://www.appbrain.com/app/finally-mature-singles-dating/com.jaumo.mature','https://www.appbrain.com/app/clue-period-tracker-calendar/com.clue.android','https://www.appbrain.com/app/womanlog-period-calendar/com.womanlog','https://www.appbrain.com/app/quit-porn-addiction-guide-app/com.quityour.pornaddiction.my_goal','https://www.appbrain.com/app/fachat-online-video-chat/com.fachat.freechat','https://www.appbrain.com/app/x-sexy-video-downloader/free.xnxx.hot.video.downloader','https://www.appbrain.com/app/english-grammar-book/com.nithra.bestenglishgrammar','https://www.appbrain.com/app/sexy-girls-live-video-call-app/mirchi.livevideo.chat','https://www.appbrain.com/app/flo-period-pregnancy-tracker/org.iggymedia.periodtracker','https://www.appbrain.com/app/meditopia-sleep-meditation/app.meditasyon','https://www.appbrain.com/app/lovely-meet-and-date-locals/com.pinkapp','https://www.appbrain.com/app/kids-police-for-parents/com.realdream.kidspolice','https://www.appbrain.com/app/live-chat-video-call-whatslive/com.download.funny.online','https://www.appbrain.com/app/parau-video-chat-with-friends/com.parau.videochat','https://www.appbrain.com/app/beauty-anime-girls-wallpapers/alpha.hd.anime.girls.wallpapers','https://www.appbrain.com/app/secret-dating-nearby-casual/ru.taboo.app','https://www.appbrain.com/app/body-editor-photo-editor/breastenlarger.bodyeditor.photoeditor','https://www.appbrain.com/app/leg-workouts-exercises-for-men/legworkout.formen.legsworkoutstraining','https://www.appbrain.com/app/butt-workout-leg-workout/buttocksworkout.legsworkout.buttandleg','https://www.appbrain.com/app/butt-leg-hips-glute-workout/legworkout.forwomen.buttocksworkout_hipsworkouts','https://www.appbrain.com/app/period-tracker-ovulation/com.brc.PeriodTrackerDiary','https://www.appbrain.com/app/lose-weight-app-for-men/menloseweight.loseweightappformen.weightlossformen','https://www.appbrain.com/app/buttocks-workout-fitness-app/buttocksworkout.hipsworkouts.forwomen.legworkout']
    avaliacao = []
    listApps = []
    driver.get('https://www.appbrain.com/app/blk-dating-meet-black-singles/com.affinityapps.blk')
    time.sleep(5)
    for url in lista:
        driver.get(url)
        time.sleep(3)
        #avaliacao = []
        #try:
        #    temp = driver.find_elements(By.CLASS_NAME, 'destinations')
        #    for aux in temp:
        #        if aux.text[0] == "P":
        #            print(aux.text)

        #except NoSuchElementException:
        #    print("")
        #except Exception as e:
        #    print("ERROR")

        #avaliacao.append(url)
        #print(avaliacao)
        #listApps.append(avaliacao)

    #for palavra in lista:
    #    partes = palavra.split('/')
    #    temp =  []
    #    # Verifica se existem pelo menos quatro ocorrências de "/"
    #    if len(partes) >= 6:
    #        # Retorna o texto após a quarta ocorrência
    #        temp.append('/'.join(partes[5:]))
    #    else:
    #        # Se não houver quatro ocorrências, retorne uma string vazia ou outra mensagem de erro, conforme necessário
    #        temp.append("")
    #    
    #    listApps.append(temp)


    #df = pd.DataFrame(listApps[1:], columns=listApps[0])
    #excel_file = "nomeApp" + ".xlsx"
    #df.to_excel(excel_file, index=False)
    #print(f"Os dados foram salvos em {excel_file}")

    #app = []
    #app.append(driver.find_element(By.XPATH, "//span[contains(@itemprop, 'description')]").text)

    #print(app[0])
    #time.sleep(1)
    #elemento = driver.find_element(By.XPATH, '/html/body/div[2]/div[3]/div[1]/div[1]/div[2]/div[2]/div[1]/span').text
    #print(driver.find_element(By.XPATH, '/html/body/div[2]/div[3]/div[1]/div[1]/div[2]/div[2]/div[1]/span').text)
    #print(driver.find_element(By.XPATH, '/html/body/div[2]/div[3]/div[1]/div[1]/div[2]/div[2]/div[2]/span').text)
    #print(driver.find_element(By.XPATH, '/html/body/div[2]/div[3]/div[1]/div[1]/div[2]/div[2]/div[3]').text)

def testesIndiv():
    service = Service()
    options = webdriver.EdgeOptions()
    driver = webdriver.Edge(service=service, options=options)
    driver.get('https://www.appbrain.com/app/wonder-ai-art-generator/com.codeway.wonder')
    #app = []
    #app.append(driver.find_element(By.XPATH, "//span[contains(@itemprop, 'description')]").text)

    #print(app[0])
    time.sleep(1)
    #elemento = driver.find_element(By.XPATH, '/html/body/div[2]/div[3]/div[1]/div[1]/div[2]/div[2]/div[1]/span').text
    #print(driver.find_element(By.XPATH, '/html/body/div[2]/div[3]/div[1]/div[1]/div[2]/div[2]/div[1]/span').text)
    #print(driver.find_element(By.XPATH, '/html/body/div[2]/div[3]/div[1]/div[1]/div[2]/div[2]/div[2]/span').text)
    print(driver.find_element(By.XPATH, '/html/body/div[2]/div[3]/div[1]/div[1]/div[2]/div[2]/div[3]').text)

def testesColetivos():
    service = Service()
    options = webdriver.EdgeOptions()
    driver = webdriver.Edge(service=service, options=options)

    driver.get('https://www.appbrain.com/app/wonder-ai-art-generator/com.codeway.wonder')
    time.sleep(2)
    header = ['appName','developerName','ratingValue','ratingCount','downloads','shortDescription','Cost of in-app purchases','Total downloads','Recent downloads','Rating','Ranking','Version','APK size',
              'Number of libraries','Designed for Android','Suitable for','Ads',]
    app = []
    app.append(driver.find_element(By.XPATH, '/html/body/div[2]/div[3]/div[1]/div[1]/div[2]/div[1]/div/div[2]/div[1]/h1').text)
    app.append(driver.find_element(By.XPATH, '/html/body/div[2]/div[3]/div[1]/div[1]/div[2]/div[1]/div/div[2]/span/a/span').text)
    app.append(driver.find_element(By.XPATH, '/html[1]/body[1]/div[2]/div[3]/div[1]/div[1]/div[2]/div[2]/div[1]/span[1]').text)
    app.append(driver.find_element(By.XPATH, '/html[1]/body[1]/div[2]/div[3]/div[1]/div[1]/div[2]/div[2]/div[2]/span[1]').text)
    app.append(driver.find_element(By.XPATH, '/html/body/div[2]/div[3]/div[1]/div[1]/div[2]/div[2]/div[3]').text)
    app.append(driver.find_element(By.XPATH, '/html/body/div[2]/div[3]/div[1]/div[1]/div[2]/h2[1]').text)
    #app.append(driver.find_element(By.XPATH, '//*[@id="main_content"]/div[7]/div[3]/table/tbody/tr[2]/td[2]').text)
    #app.append(driver.find_element(By.XPATH, '//*[@id="main_content"]/div[7]/div[3]/table/tbody/tr[3]/td[2]').text)
    #app.append(driver.find_element(By.XPATH, '//*[@id="main_content"]/div[7]/div[3]/table/tbody/tr[4]/td[2]').text)
    #app.append(driver.find_element(By.XPATH, '//*[@id="main_content"]/div[7]/div[3]/table/tbody/tr[5]/td[2]/a').text)
    #app.append(driver.find_element(By.XPATH, '//*[@id="main_content"]/div[7]/div[3]/table/tbody/tr[6]/td[2]/a').text)
    #app.append(driver.find_element(By.XPATH, '//*[@id="main_content"]/div[7]/div[3]/table/tbody/tr[7]/td[2]').text)
    #app.append(driver.find_element(By.XPATH, '//*[@id="main_content"]/div[7]/div[3]/table/tbody/tr[8]/td[2]').text)
    #app.append(driver.find_element(By.XPATH, '//*[@id="main_content"]/div[7]/div[3]/table/tbody/tr[9]/td[2]/a').text)
    #app.append(driver.find_element(By.XPATH, '//*[@id="main_content"]/div[7]/div[3]/table/tbody/tr[10]/td[2]').text)
    #app.append(driver.find_element(By.XPATH, '//*[@id="main_content"]/div[7]/div[3]/table/tbody/tr[11]/td[2]').text)
    #app.append(driver.find_element(By.XPATH, '//*[@id="main_content"]/div[7]/div[3]/table/tbody/tr[12]/td[2]').text)
    #app.append(driver.find_element(By.XPATH, '').text)
    #app.append(driver.find_element(By.XPATH, '').text)
    #app.append(driver.find_element(By.XPATH, '').text)
    #app.append(driver.find_element(By.XPATH, '').text)
    #app.append(driver.find_element(By.XPATH, '').text)
    #app.append(driver.find_element(By.XPATH, '').text)
    #app.append(driver.find_element(By.XPATH, '').text)
    #app.append(driver.find_element(By.XPATH, '').text)
    #app.append(driver.find_element(By.XPATH, '').text)
    #app.append(driver.find_element(By.XPATH, '').text)
    #app.append(driver.find_element(By.XPATH, '').text)
    #app.append(driver.find_element(By.XPATH, '').text)
    #app.append(driver.find_element(By.XPATH, '/html/body/div[2]/div[3]/div[1]/div[1]/div[2]/div[7]/div[3]/table/tbody/tr[11]/td[2]').text)
    
    
    #print(header)
    print(app)

def tentativaAtual():
    service = Service()
    options = webdriver.EdgeOptions()
    driver = webdriver.Edge(service=service, options=options)
    apps = ["https://www.appbrain.com/app/wonder-ai-art-generator/com.codeway.wonder"]

    listApps = []
    #listApps.append(["appName", "appDev", "appDesc", "appSco", "appPrice", "appDow", "appL", "appCat", "appLink"])

    contGeral = 0

    for i in range(1):

        url = apps[0]

        driver.get(url)

        #eventos clickers
        #driver.find_element(By.ID, 'descLink').click()
        print("Check")
        #coleta de dados em sí
        appName = driver.find_element(By.CLASS_NAME, 'app-top-title')
        #appFullDescription = driver.find_element(By.ID, 'descContents')
        appPlayStoreLink = driver.find_element(By.CSS_SELECTOR, ".btn btn-primary [href]")
        
        #appRating = driver.find_element(By.CLASS_NAME, 'text-center pe-3 right-border')
        #appReviewNumber = driver.find_element(By.CLASS_NAME, 'text-center px-3 right-border')
        appTableAbout = driver.find_elements(By.CLASS_NAME, 'table table-striped')
        
        driver.find_element(By.XPATH, '/html/body/div[2]/div[3]/div[1]/div[1]/div[2]/div[1]/div/div[2]/span/a/span').text


        #appTotalDownloads = driver.find_element(By.CLASS_NAME, 'text-center px-3 right-border')

        #appSco = driver.find_elements(By.CLASS_NAME, 'browse-app-large-score')
        #appMor = driver.find_elements(By.CLASS_NAME, 'browse-app-large-bottom-item')
        
        #appLik = driver.find_elements(By.CLASS_NAME, 'vmargin-s')
        #appLik = driver.find_elements_by_css_selector(".vmargin-s [href]")
        #appLik = driver.find_elements(By.CSS_SELECTOR, ".vmargin-s [href]")
        
        listTemp = []

        listTemp.append(appName.text)
        #listTemp.append(appFullDescription.text)
        #listTemp.append(appPlayStoreLink.get_attribute("href"))
        #listTemp.append(appRating.text)
        #listTemp.append(appReviewNumber.text)

        for texto in appPlayStoreLink:
            listTemp.append(texto.text)
            
        #morCount = 0
        #likCount = 0
        #for x in range(20):
        #    listTemp = []
        #    listTemp.append(appNam[x].text)
        #    print(str(contGeral) + ' - ' + str(likCount) + ' - ' + str(x)+ ' - ' + str(morCount) + ' - ' + categoria + " - " + appNam[x].text)
        #    listTemp.append(appDev[x].text)
        #    listTemp.append(appDes[x].text)
        #    listTemp.append(appSco[x].text)
        #    listTemp.append(appMor[morCount].text)
        #    listTemp.append(appMor[morCount+1].text)
        #    listTemp.append(appMor[morCount+2].text)
        #    listTemp.append(appMor[morCount+3].text)
#
        #    if appNam[x].text == '':
        #        likCount = likCount + 1
        #    listTemp.append(appLik[likCount].get_attribute("href"))
        #    
        #    likCount = likCount + 2
        #    morCount = morCount + 4
#
        #    contGeral = contGeral + 1
            
        listApps.append(listTemp)
                

    for a in listApps:
        print(a)
            
    #df = pd.DataFrame(listApps[1:], columns=listApps[0])
    #excel_file = categoria + ".xlsx"
    #df.to_excel(excel_file, index=False)
    #print(f"Os dados foram salvos em {excel_file}")

def confianteDeMais():
    service = Service()
    options = webdriver.EdgeOptions()
    driver = webdriver.Edge(service=service, options=options)
    categorias = ["art-and-design","auto-and-vehicles","beauty"]
                  #,"books-and-reference","business","comics","communication","dating",
                  #"education","entertainment","events","finance","food-and-drink","health-and-fitness","house-and-home",
                  #"libraries-and-demo","lifestyle","maps-and-navigation","medical","music-and-audio","news-and-magazines",
                  #"parenting","personalization","photography","productivity","shopping","social","sports","tools",
                  #"travel-and-local","video-players-and-editors","weather"]

    listApps = []
    listApps.append(["appName", "appDev", "appDesc", "appSco", "appPrice", "appDow", "appL", "appCat", "appLink"])
    morCountPages = 0
    contGeral1 = 0
    for categoria  in categorias: 
        morCountPages = 0
        contGeral2 = 0
        for i in range(10):
            url = 'https://www.appbrain.com/apps/most-downloaded/'+ categoria +'/free/?o=' + str(morCountPages)
            morCountPages = morCountPages + 20

            driver.get(url)


            #appName          = driver.find_elements(By.CLASS_NAME, 'browse-app-large-title')
            #appDescription  = driver.find_element(By.ID, 'descContents').text
            #appDeveloper    = driver.find_element(By.XPATH, '//*[@id="main_content"]/div[1]/div/div[2]/span').text
            #appAbout        = driver.find_elements(By.CLASS_NAME, 'table table-striped')

            appNam = driver.find_elements(By.CLASS_NAME, 'browse-app-large-title')
            appDev = driver.find_elements(By.CLASS_NAME, 'browse-app-large-author')
            appDes = driver.find_elements(By.CLASS_NAME, 'browse-app-large-description')
            appSco = driver.find_elements(By.CLASS_NAME, 'browse-app-large-score')
            appMor = driver.find_elements(By.CLASS_NAME, 'browse-app-large-bottom-item')
            
            #appLik = driver.find_elements(By.CLASS_NAME, 'vmargin-s')
            #appLik = driver.find_elements_by_css_selector(".vmargin-s [href]")
            appLik = driver.find_elements(By.CSS_SELECTOR, ".vmargin-s [href]")
            

            morCount = 0
            likCount = 0
            for x in range(20):
                listTemp = []
                listTemp.append(contGeral1)
                listTemp.append(appNam[x].text)
                print(str(contGeral1) + ' - ' + str(contGeral2) + ' - ' + str(likCount) + ' - ' + str(x)+ ' - ' + str(morCount) + ' - ' + categoria + " - " + appNam[x].text)
                listTemp.append(appDev[x].text)
                listTemp.append(appDes[x].text)
                listTemp.append(appSco[x].text)
                listTemp.append(appMor[morCount].text)
                listTemp.append(appMor[morCount+1].text)
                listTemp.append(appMor[morCount+2].text)
                listTemp.append(appMor[morCount+3].text)

                if appNam[x].text == '':
                    likCount = likCount + 1
                listTemp.append(appLik[likCount].get_attribute("href"))
                
                likCount = likCount + 2
                morCount = morCount + 4

                contGeral2 = contGeral2 + 1
                contGeral1 = contGeral1 + 1
                listApps.append(listTemp)
                    

    #for a in listApps:
    #    print(a)
            
    df = pd.DataFrame(listApps[1:], columns=listApps[0])
    excel_file = "oll.xlsx"
    df.to_excel(excel_file, index=False)
    print(f"Os dados foram salvos em {excel_file}")

def comEdge3():
    print("Oi")
    #if appNam[x].text == '':
    #listTemp.append(appLik[x].find_element(By.CLASS_NAME, 'browse-app-large-title').text)
    #listTemp.append(appLik[x].find_element(By.CLASS_NAME, 'browse-app-large-author').text)
    #listTemp.append(appLik[x].find_element(By.CLASS_NAME, 'browse-app-large-description').text)
    #listTemp.append(appLik[x].find_element(By.CLASS_NAME, 'browse-app-large-score').text)
    #temp = appLik[x].find_elements(By.CLASS_NAME, 'browse-app-large-bottom-item')
    #listTemp.append(temp[0].text)
    #listTemp.append(temp[1].text)
    #listTemp.append(temp[2].text)
    #listTemp.append(temp[3].text)
    #likCount = likCount + 2
    #
    #else: 
    #listTemp.append(appNam[x].text)
    #listTemp.append(appDev[x].text)
    #listTemp.append(appDes[x].text)
    #listTemp.append(appSco[x].text)
    #listTemp.append(appMor[morCount].text)
    #listTemp.append(appMor[morCount+1].text)
    #listTemp.append(appMor[morCount+2].text)
    #listTemp.append(appMor[morCount+3].text)

    #if appNam[x].text == '':
    #    likCount = likCount + 2
            

def comEdge():
    #service = Service()
    #options = webdriver.EdgeOptions()
    driver = webdriver.Edge(r'C:\\msedgedriver.exe')

    url = 'https://www.appbrain.com/app/canva-design-photo-video/com.canva.editor'

    driver.get(url)

    #clicka nos elementos para carregar as informações dinâmicas da tela.
    #driver.find_element(By.ID, 'descLink').click()
    #driver.find_element(By.CLASS_NAME, 'link').click()
    #driver.execute('clickShowTech()')
    appName         = driver.find_element(By.CLASS_NAME, 'app-top-title').text
    #appDescription  = driver.find_element(By.ID, 'descContents').text
    appDeveloper    = driver.find_element(By.XPATH, '//*[@id="main_content"]/div[1]/div/div[2]/span').text
    appAbout        = driver.find_elements(By.CLASS_NAME, 'table table-striped')


    print("\n\n\n\n\n\n\n")
    print(appName)
    #print(appDescription)
    #print(appDeveloper)
    print(appAbout.text)




def usingRequests():
    link = 'https://www.appbrain.com/app/toilet-monster-survival-quest/com.osgame.skibiditoilet.survival'
    header = {"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36 Edg/118.0.2088.76"}
    response = requests.get(link, headers=header)

    #response.
    
    site = BeautifulSoup(response.text, "html.parser")

    print(site.prettify())

def usingRequestsHTML():

    #session = HTMLSession()
    link = 'https://www.appbrain.com/app/wonder-ai-art-generator/com.codeway.wonder'
    header = {"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36 Edg/118.0.2088.76"}
    response = requests.get(link, headers=header)

    #response.
    
    site = BeautifulSoup(response.text, "html.parser")

    print(site.prettify())

def docSelenium():
    driver = webdriver.Chrome()
    driver.get('https://www.appbrain.com/app/canva-design-photo-video/com.canva.editor')

    description = driver.find_element(By.ID, "descContents")

def usingSlenium():
    ## Configure o caminho para o ChromeDriver aqui
    #chrome_driver_path = 'C:\Users\jonat\OneDrive\Documentos\ProjetosPython\chromedriver.exe'
    #
    ## Configurar opções do Chrome (opcional)
    #chrome_options = webdriver.ChromeOptions()
    ## Adicionar opções, se necessário, como headless, proxy, etc.
    #
    ## Crie o driver do Chrome
    #driver = webdriver.Chrome(executable_path=chrome_driver_path, chrome_options=chrome_options)


    driver = webdriver.Chrome()


    driver.get('https://www.appbrain.com/app/canva-design-photo-video/com.canva.editor')

    html = driver.page_source
    driver.quit()

    print(html)

    #soup = BeautifulSoup(html.content, "div#descContents")
    #soup = BeautifulSoup(html)
    #
    #print(soup.get_text())
    #
    ## check out the docs for the kinds of things you can do with 'find_all'
    ## this (untested) snippet should find tags with a specific class ID
    ## see: http://www.crummy.com/software/BeautifulSoup/bs4/doc/#searching-by-css-class
    #for tag in soup.find_all("a", class_="my_class"):
    #    print tag.text


#usingRequests()
#usingSlenium()
#docSelenium()
#vamoBR()
#tentativaAtual()
#testesIndiv()
#semZeradosPfv()
#testesColetivos()
#confianteDeMais()
#comEdge()
#beatifulSoupSelenium()
#listaInfo()
bot_lindinho_2009()
#beatifulSoupSeleniumWire()
#usingRequestsHTML()